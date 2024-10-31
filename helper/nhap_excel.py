from helper.utils import connect_db, execute_query, close_db, execute_query_data
from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import pandas as pd
import numpy as np

def getConditionQuery(filters):
    conditions = []
    for key, value in filters.items():
        if value.get('value'):
            match value.get("type"):
                case "approximately":
                    conditions.append(f"{key} LIKE '%{value.get('value')}%'")
                case "gte":
                    conditions.append(f"{key} >= '{value.get('value')}'")
                case "lte":
                    conditions.append(f"{key} <= '{value.get('value')}'")
                case _: 
                    conditions.append(f"{key} = '{value.get('value')}'")
    if conditions:
        query = " WHERE " + " AND ".join(conditions)
    else:
        query = ""
    return query

def get_data(filters, page, size, table, order_by):
    try: 
        conn = connect_db()
        offset = (page - 1) * size + 1
        query = f"SELECT *, ROW_NUMBER() OVER (ORDER BY {order_by}) AS RowNum FROM {table}"

        conditionQuery = getConditionQuery(filters)
        query += conditionQuery

        last_query = f"WITH TEMP AS ({query}) SELECT * FROM TEMP WHERE RowNum BETWEEN {offset} AND {offset + size - 1}"

        cursor = execute_query(conn, last_query)
        rows = cursor.fetchall()

        count_query = f"SELECT COUNT(*) FROM {table}"
        count_query += conditionQuery
        cursor2 = execute_query(conn, count_query)
        total = cursor2.fetchall()[0][0]
        close_db(conn)
        return {
            "data": rows,
            "total": total
        }
    except Exception as e:
        print(e)
        return {
            "data": [],
            "total": 0
        }

def get_excel_from_table(database, table, filename, filters, dateCols = []):
    try:
        conn = connect_db()
        query_header = f"""SELECT COLUMN_NAME
                FROM [{database}].INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = '{table}';"""

        cursor = execute_query(conn, query_header)
        rows = cursor.fetchall()
        headers = [row[0] for row in rows]
        lower_headers = [h.lower() for h in headers]
        date_indexs = [lower_headers.index(col) for col in dateCols if col in lower_headers]

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws.append(headers)

        query = f"SELECT * FROM [{database}].[dbo].[{table}]"
        conditionQuery = getConditionQuery(filters)
        query += conditionQuery

        cursor = execute_query(conn, query)
        data = cursor.fetchall()

        for row in data:
            ws.append(list(row))
            
        date_format = NamedStyle(name="date", number_format="DD/MM/YYYY")

        thin_border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
        
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for i, col in enumerate(ws.columns, 0):
            max_length = 0
            column = get_column_letter(i + 1)
            for cell in col:
                try:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if i in date_indexs:
                        cell.value = datetime.strptime(str(cell.value), "%Y-%m-%d") if cell.value else cell.value       
                        cell.style = date_format   
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[column].width = max(adjusted_width,7)
        

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        current_time = datetime.now().strftime("%d/%m/%Y_%H_%M_%S")
        close_db(conn)
        
        response = make_response(buffer.read())
        response.headers.set('Content-Disposition', f'attachment; filename="{filename}{current_time}.xlsx"')
        response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return response    
    except Exception as e:
        print(e)
        return None
    
def upload_excel_to_db(database, table, file):
    try:
        df = pd.read_excel(file)
        df = df.applymap(lambda x: None if pd.isna(x) else x)
        data_tuples = [tuple(row) for row in df.itertuples(index=False, name=None)]

        if len(data_tuples) > 0:
            conn = connect_db()
            len_row = len(data_tuples[0])
            placeholder = ",".join(["?" for _ in range(len_row)])
            query = f"INSERT INTO [{database}].[dbo].[{table}] VALUES ({placeholder})"
            execute_query_data(conn, query, data_tuples)
            conn.commit()
            close_db(conn)
    except Exception as e:
        print(e)
        return None