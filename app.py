from flask import render_template, request, url_for, redirect, g, flash, jsonify, send_file, session, flash, get_flashed_messages, render_template_string,make_response
from flask_paginate import Pagination, get_page_parameter
from flask_login import LoginManager, UserMixin, login_user, logout_user, current_user, login_required
import datetime
from functools import wraps
from pandas import DataFrame,read_excel,ExcelWriter,to_numeric,to_datetime
from openpyxl import load_workbook
import os
import time
from io import BytesIO
import subprocess
import numpy as np
from waitress import serve
import sys
from openpyxl.styles import Font, PatternFill, NamedStyle
from config_app import *
from helper.utils import *

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class Nhanvien(UserMixin, db.Model):
    __tablename__ = 'Nhanvien'
    id = db.Column(db.Integer, primary_key=True)
    macongty = db.Column(db.String(10), nullable=False)
    masothe = db.Column(db.Integer, nullable=False)
    hoten = db.Column(db.Unicode(50), nullable=False)
    phongban = db.Column(db.String(10), nullable=False)
    capbac = db.Column(db.String(10), nullable=False)
    phanquyen = db.Column(db.String(10), nullable=False)
    matkhau = db.Column(db.String(10), nullable=False)

    def __repr__(self):
        return f"<User {self.hoten}>"

def chuyen_so_thanh_sotien(so):
    return "{:,.0f}".format(so)

def get_line(masothe,macongty):
    try:
        conn = connect_db()
        query = f"select CHUYEN from [INCENTIVE].[dbo].[DS_TO_TRUONG] where MST='{masothe}' and NHA_MAY='{macongty}'"
        # 
        cursor = execute_query(conn, query)
        rows = cursor.fetchall()
        result = [row[0] for row in rows]
        # print(result)
        close_db(conn)
        return result
    except:
        return []
    
def get_all_styles(ngay, chuyen):
    try:
        if ngay and chuyen:
            conn = connect_db()
            query = f"SELECT Distinct STYLE FROM [INCENTIVE].[dbo].[SL_CA_NHAN] WHERE NGAY='{ngay}' AND CHUYEN='{chuyen}'"
            # 
            cursor = execute_query(conn, query)
            result = cursor.fetchall()
            close_db(conn)
            return [style[0] for style in result]
        else:
            return []
    except:
        return []

def laydulieuthuongmaychitiet():
    conn = connect_db()
    query = f"SELECT * FROM [INCENTIVE].[dbo].[THUONG_CN_MAY_HANG_NGAY_CHI_TIET] order by Ngay desc, Chuyen Asc, Cast(MST as int) Asc"
    cursor = execute_query(conn, query) 
    result = cursor.fetchall()
    close_db(conn)
    return result

def lay_danhsach_congnhan_trongchuyen(chuyen):
    conn = connect_db()
    query = f"SELECT * FROM [INCENTIVE].[dbo].[DS_CN_MAY_THEO_HC_CHUYEN] WHERE Chuyen='{chuyen}'"
    cursor = execute_query(conn, query) 
    result = cursor.fetchall()
    close_db(conn)
    return list(result)
    
def lay_danhsach_chuyen_hotro(chuyen):
    try:
        if chuyen:
            conn = connect_db()
            if "S"in chuyen and len(chuyen)==5:
                query = f"SELECT * FROM [INCENTIVE].[dbo].[DS_CHUYEN_MAY] WHERE LINE LIKE '{chuyen[0]}_S__' ORDER BY LINE"
            else:
                query = f"SELECT * FROM [INCENTIVE].[dbo].[DS_CHUYEN_MAY] WHERE LINE LIKE '{chuyen[0]}%' ORDER BY LINE"
            cursor = execute_query(conn, query) 
            result = cursor.fetchall()
            close_db(conn)
            return [line[0] for line in result]
        else:
            return []
    except:
        return []
  
def lay_danhsach_sanluong(ngay, chuyen, style,mst,hoten,macongdoan):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[SL_CA_NHAN] WHERE 1=1 "
        if ngay:
            query += f"AND NGAY='{ngay}' "
        else:
            return []
        if chuyen:
            query += f"AND CHUYEN='{chuyen}' "
        if style:
            query += f"AND STYLE='{style}' "
        else:
            return []
        if mst:
            query += f"AND MST='{mst}' "
        if hoten:
            query += f"AND HO_TEN=N'{hoten}' "
        if macongdoan:
            query += f"AND MA_CONG_DOAN ='{macongdoan}' "
        query += "ORDER BY CAST(MST as INT) ASC, MA_CONG_DOAN ASC"
        
        cursor = execute_query(conn, query) 
        result = cursor.fetchall()
        close_db(conn)
        return list(result)
    except Exception as e:
        print(e)
        return []

def capnhat_sanluong(mst,hoten,chuyen,ngay,style,macongdoan,sanluong):
    conn = connect_db()
    query = f"INSERT INTO [INCENTIVE].[dbo].[SL_CA_NHAN] (MST,HO_TEN,CHUYEN,NGAY,STYLE,MA_CONG_DOAN,SL_CA_NHAN) VALUES('{mst}', N'{hoten}', '{chuyen}', '{ngay}', '{style}', '{macongdoan}', '{sanluong}')"
    # print(query)
    try:   
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return {"ketqua":True}
    except Exception as e:
        
        return {
                "ketqua":False,
                "lido":e,
                "query": query}

def xoa_sanluong(id):
    conn = connect_db()
    query = f"DELETE FROM [INCENTIVE].[dbo].[SL_CA_NHAN] WHERE ID='{id}'"
    execute_query(conn, query)
    try:
        conn.commit()
        close_db(conn)
        return True
    except Exception as e:
        print(e)
        return False

def lay_tencongdoan(thongtin):
    try:
        macongdoan = thongtin.split("_")[0]
        style = thongtin.split("_")[1]
        conn = connect_db()
        cursor = execute_query(conn, f"SELECT TEN_CONG_DOAN FROM [INCENTIVE].[dbo].[SAM_SEW] WHERE STYLE='{style}' AND MA_CONG_DOAN='{macongdoan}'")
        result = cursor.fetchone()
        close_db(conn)
        return result[0]
    except:
        return None

def them_nguoi_di_hotro(nhamay,chuyen,mst,hoten,chucdanh,chuyendihotro,ngaydieuchuyendi,giodieuchuyendi,sogiohotro):
    try:
        conn = connect_db()
        if giodieuchuyendi:
            query = f"insert into [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] values ('{nhamay}','{mst}',N'{hoten}',N'{chucdanh}','{chuyen}','{chuyendihotro}','{ngaydieuchuyendi}','{giodieuchuyendi}','{sogiohotro}')"
        else:
            query = f"insert into [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] values ('{nhamay}','{mst}',N'{hoten}',N'{chucdanh}','{chuyen}','{chuyendihotro}','{ngaydieuchuyendi}',NULL,'{sogiohotro}')"
        # 
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except:
        return False

def lay_danhsach_di_hotro(chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] WHERE CHUYEN='{chuyen}'"
        
        cursor = execute_query(conn, query)
        result = cursor.fetchall()
        close_db(conn)
        return list(result)
    except:
        return []
    
def lay_danhsach_den_hotro(chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] WHERE CHUYEN_DI_HO_TRO ='{chuyen}'"
        
        cursor = execute_query(conn, query)
        result = cursor.fetchall()
        close_db(conn)
        return list(result)
    except:
        return []

def lay_danhsach_di_hotro_tatca(mst,chuyen,ngay):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] WHERE 1=1 "
        if mst:
            query +=f" AND MST='{mst}'"
        if chuyen:
            query +=f" AND CHUYEN='{chuyen}'"
        if ngay:
            query +=f" AND NGAY_DI_HO_TRO='{ngay}'"
        query += " ORDER BY NGAY_DI_HO_TRO DESC, GIO_DI_HO_TRO ASC"
        cursor = execute_query(conn, query)
        result = cursor.fetchall()
        close_db(conn)
        return list(result)
    except:
        return []

def lay_danhsach_tnc_chua_lenchuyen(chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[CN_TNC_CHUA_NGOI_CHUYEN] WHERE CHUYEN LIKE '{chuyen[0]}TNC%' ORDER BY CAST(MST as INT) ASC"
        cursor = execute_query(conn, query)
        result = cursor.fetchall()
        close_db(conn)
        return list(result)
    except:
        return []
    
def lay_danhsach_tnc_ngoichuyen(chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[CN_TNC_NGOI_CHUYEN] WHERE CHUYEN_NGOI_LV = '{chuyen}' ORDER BY CAST(MST as INT) ASC"
        
        cursor = execute_query(conn, query)
        result = cursor.fetchall()
        close_db(conn)
        return list(result)
    except:
        return []

def thaydoi_tungay_cn_tnc(id,ngay):
    try:
        conn = connect_db()
        query = f"update [INCENTIVE].[dbo].[CN_TNC_NGOI_CHUYEN] SET TU_NGAY='{ngay}' WHERE ID='{id}'"
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except:
        return False
    
def thaydoi_denngay_cn_tnc(id,ngay):
    try:
        conn = connect_db()
        query = f"update [INCENTIVE].[dbo].[CN_TNC_NGOI_CHUYEN] SET DEN_NGAY='{ngay}' WHERE ID='{id}'"
        print(query)
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except:
        return False

def nhan_tnc_len_chuyen(id,chuyen):
    try:
        conn = connect_db()
        query = f"update [INCENTIVE].[dbo].[CN_TNC_NGOI_CHUYEN] SET CHUYEN_NGOI_LV='{chuyen.replace('[','').replace(']','').replace("'","")}' WHERE ID='{id}'"
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except:
        return False
    
def capnhat_sogio_hotro(id,sogio):
    try:
        conn = connect_db()
        query = f"update [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] SET SO_GIO_HO_TRO='{sogio}' WHERE ID='{id}'"
        # 
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except:
        return False
        
def capnhat_sogio_den_hotro(id,sogio):
    try:
        conn = connect_db()
        query = f"update [INCENTIVE].[dbo].[CN_MAY_DI_HO_TRO] SET SO_GIO_HO_TRO='{sogio}' WHERE ID='{id}'"
        # 
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except:
        return False
    
def laytongsanluongtheocongdoan(ngay,chuyen,style,macongdoan):
    try:
        conn = connect_db()
        if macongdoan:
            query = f"select MA_CONG_DOAN,QTY from [INCENTIVE].[dbo].[TONG_SL_CONG_DOAN] where NGAY='{ngay}' and CHUYEN='{chuyen}' and STYLE='{style}' and MA_CONG_DOAN='{macongdoan}' group by MA_CONG_DOAN,QTY"
        else:
            query = f"select MA_CONG_DOAN,QTY from [INCENTIVE].[dbo].[TONG_SL_CONG_DOAN] where NGAY='{ngay}' and CHUYEN='{chuyen}' and STYLE='{style}' group by MA_CONG_DOAN,QTY"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        result=[]
        for row in rows:
            if row[0]:
                result.append([row[0],row[1]])
        return result
    except:
        return []
    
def lay_sanluong_tong_theochuyen(ngay, chuyen, style):
    try:
        if ngay and chuyen and style:
            conn = connect_db()
            query = f"select QTY from [INCENTIVE].[dbo].[SL_NGAY_CHUYEN_STYLE ] where NGAY='{ngay}' and CHUYEN='{chuyen}' and GR_STYLE='{style}'"
            # 
            result = execute_query(conn, query).fetchone()
            close_db(conn)
            return result[0]
        else:
            return 0
    except:
        return 0

def lay_baocao_thuong_congnhan_may(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT MST,HO_TEN,CHUYEN,NGAY,SAH,SCP,SO_GIO,Eff_CA_NHAN,THUONG_CA_NHAN FROM [INCENTIVE].[dbo].[INCENTIVE_CN_MAY_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        print
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_cat(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT MST,HO_TEN,CHUYEN,NGAY,HE_SO,TGLV,HE_SO_CN,TONG_HE_SO,THUONG_NHOM,THUONG_CN FROM [INCENTIVE].[dbo].[INCENTIVE_CN_CAT_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_la(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_CN_LA_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []

def lay_baocao_thuong_congnhan_qc1(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_CN_QC1_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, NHOM ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_qc2(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_CN_QC2_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, NHOM ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_donggoi(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_CN_DONG_GOI_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_ndc(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_CN_NDC_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, NHOM ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_phu(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_CN_PHU_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_quanly(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[INCENTIVE_QUANLY_HANG_NGAY] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except:
        return []
    
def lay_baocao_thuong_congnhan_nhommay(macongty,ngay,chuyen,style):
    try:
        conn = connect_db()
        query = f"SELECT Workdate,Line,Sah,Total_hours,Eff,Style,TRANG_THAI,CHUYEN_MOI,OQL,GR_INCENTIVE,GR_INCENTIVE_TOPUP1,GR_INCENTIVE_TOPUP2,TONG_THUONG FROM [INCENTIVE].[dbo].[THUONG_NHOM_MAY_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND Line LIKE '{macongty}%'"
        if ngay:
            query += f" AND Workdate='{ngay}'"
        if chuyen:
            query += f" AND Line LIKE '%{chuyen}%'" 
        if style:
            query += f" AND Style LIKE '%{style}%'"
        query += "ORDER BY Workdate DESC, Line ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_thuong_congnhan_nhomcat(macongty,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT NGAY,NHOM,SAH,SO_GIO,EFF,CAT_KE,NHA_MAY,UI,BE,BE_TOPUP1,TOP_UP1,TI_LE_LOI,AQL,THUONG_NHOM,KICH_CAU1,TONG_THUONG_NHOM FROM [INCENTIVE].[dbo].[THUONG_NHOM_CAT_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'" 
        query += "ORDER BY NGAY DESC, NHOM ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_thuong_congnhan_nhomla(macongty,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[THUONG_NHOM_LA_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'" 
        query += "ORDER BY NGAY DESC, CHUYEN ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_thuong_congnhan_nhomdonggoi(macongty,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[THUONG_NHOM_DONG_GOI_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'" 
        query += "ORDER BY NGAY DESC, NHOM ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_thuong_congnhan_nhomndc(macongty,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[THUONG_NHOM_NDC_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'" 
        query += "ORDER BY NGAY DESC, NHOM ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_thuong_congnhan_nhomqc1(macongty,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[THUONG_NHOM_QC1_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'" 
        query += "ORDER BY NGAY DESC, NHOM ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_thuong_congnhan_nhomqc2(macongty,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[THUONG_NHOM_QC2_HANG_NGAY] WHERE 1=1"
        if macongty:
            query += f" AND NHOM LIKE '{macongty}%'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND NHOM LIKE '%{chuyen}%'" 
        query += "ORDER BY NGAY DESC, NHOM ASC"
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_sogio_lamviec(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT MST,HO_TEN,CHUYEN,NGAY,SO_GIO,CHUC_VU FROM [INCENTIVE].[dbo].[TGLV_1] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_sanluong_canhan(macongty,mst,ngay,chuyen):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[SL_CA_NHAN_1] WHERE 1=1" 
        if macongty:
            query += f" AND CHUYEN LIKE '{macongty}%'"
        if mst:
            query += f" AND MST='{mst}'"
        if ngay:
            query += f" AND NGAY='{ngay}'"
        if chuyen:
            query += f" AND CHUYEN LIKE '%{chuyen}%'"
        query += " ORDER BY NGAY DESC, CHUYEN ASC, MST ASC"
        # 
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
    
def lay_baocao_scp_canhan(macongty,mst,tungay,denngay):
    try:
        conn = connect_db()
        query = f"SELECT * FROM [INCENTIVE].[dbo].[SCP_CA_NHAN] WHERE 1=1" 
        if macongty:
            query += f" AND NHA_MAY = 'NT{macongty}'"
        if mst:
            query += f" AND MST = '{mst}'"
        if tungay:
            query += f" AND TU_NGAY <= '{tungay}'"
        if denngay:
            query += f" AND DEN_NGAY >= '%{denngay}%'"
        query += " ORDER BY CAST(MST As INT) ASC"
        
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []

def capnhat_danhsach_dihotro(chuyendihotro,ngay,gio,sogio,id):
    try:
        conn = connect_db()
        query = f"""
            update INCENTIVE.dbo.CN_MAY_DI_HO_TRO
            set CHUYEN_DI_HO_TRO='{chuyendihotro}',NGAY_DI_HO_TRO='{ngay}',GIO_DI_HO_TRO='{gio}',SO_GIO_HO_TRO='{sogio}'
            where ID='{id}'
        """
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except Exception as e:
        print(e)
        return False

def doichuyendihotro(chuyen,id):
    try:
        conn = connect_db()
        query = f"""
            update INCENTIVE.dbo.CN_MAY_DI_HO_TRO
            set CHUYEN_DI_HO_TRO='{chuyen}'
            where ID='{id}'
        """
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except Exception as e:
        print(e)
        return False
        
def doingaydihotro(ngay,id):
    try:
        conn = connect_db()
        query = f"""
            update INCENTIVE.dbo.CN_MAY_DI_HO_TRO
            set NGAY_DI_HO_TRO='{ngay}'
            where ID='{id}'
        """
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except Exception as e:
        print(e)
        return False
        
def doigiodihotro(gio,id):
    try:
        conn = connect_db()
        query = f"""
            update INCENTIVE.dbo.CN_MAY_DI_HO_TRO
            set GIO_DI_HO_TRO='{gio}'
            where ID='{id}'
        """
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except Exception as e:
        print(e)
        return False
        
def doisogiodihotro(sogio,id):
    try:
        conn = connect_db()
        query = f"""
            update INCENTIVE.dbo.CN_MAY_DI_HO_TRO
            set SO_GIO_HO_TRO='{sogio}'
            where ID='{id}'
        """
        
        execute_query(conn, query)
        conn.commit()
        close_db(conn)
        return True
    except Exception as e:
        print(e)
        return False

def lay_baocao_hieusuat_chitiet_may(nam,thang,macongty,mst,chuyen):
    try:
        conn = connect_db()
        query = f"""select * from [INCENTIVE].[dbo].[TONG_HOP_EFF_MST_CHUYEN_LV] Where 1=1 """
        if not thang:
            thang = datetime.datetime.now().month
        if not nam:
            nam =  datetime.datetime.now().year
        query += f" and Thang={thang} and Nam={nam}"
        if macongty:
            query += f" and CHUYEN like '{macongty}%'"
        if mst:
            query += f" and MST = '{thang}' "
        if chuyen:
            query += f" and CHUYEN = '{chuyen}' "
        query += " order by NAM desc, THANG desc, CHUYEN asc, MST asc"
        # print(query)
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []
        
def lay_baocao_hieusuat_tonghop_may(nam,thang,macongty,mst,chuyen):
    try:
        conn = connect_db()
        query = f"""select * from [INCENTIVE].[dbo].[TONG_HOP_EFF_CN_MAY] Where 1=1 """
        if not thang:
            thang = datetime.datetime.now().month
        if not nam:
            nam =  datetime.datetime.now().year
        query += f" and Thang={thang} and Nam={nam}"
        if macongty:
            query += f" and CHUYEN like '{macongty}%'"
        if mst:
            query += f" and MST = '{thang}' "
        if chuyen:
            query += f" and CHUYEN = '{chuyen}' "
        query += " order by NAM desc, THANG desc, CHUYEN asc, MST asc"
        # print(query)
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []

def lay_baocao_tienthuong_may(nam,thang,macongty,mst,chuyen):
    try:
        conn = connect_db()
        query = f"""select * from [INCENTIVE].[dbo].[TONG_HOP_TIEN_THUONG_CN_MAY] Where 1=1 """
        if not thang:
            thang = datetime.datetime.now().month
        if not nam:
            nam =  datetime.datetime.now().year
        query += f" and Thang={thang} and Nam={nam}"
        if macongty:
            query += f" and CHUYEN like '{macongty}%'"
        if mst:
            query += f" and MST = '{thang}' "
        if chuyen:
            query += f" and CHUYEN = '{chuyen}' "
        query += " order by NAM desc, THANG desc, CHUYEN asc, MST asc"
        # print(query)
        rows = execute_query(conn, query).fetchall()
        close_db(conn)
        return rows
    except Exception as e:
        print(e)
        return []

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.is_anonymous:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def before_request():
    try:
        if current_user.is_authenticated:
            lines = get_line(current_user.masothe, current_user.macongty)
            # print(lines)
            if lines:
                if len(lines) == 1:
                    g.notice = {"line":lines,"role":"tt"}
                elif len(lines) > 1:
                    g.notice = {"line":lines,"role":"tk"}
            else:
                g.notice = {"line":[],"role":""}
        else:
            g.notice = {"line":[],"role":""}
    except:
        g.notice = {"line":[],"role":""}
        
@app.context_processor
def inject_notice():
    return dict(notice=getattr(g, 'notice', {}))  
    
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Nhanvien, user_id)

@app.errorhandler(404)
def page_not_found(e):
    return render_template_string("Trang không tìm thấy, vui lòng <a href='/'>quay lại</a> trang chủ"), 404

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        try:
            macongty = request.form['macongty']
            masothe = request.form['masothe']
            matkhau = request.form['matkhau']
            user = Nhanvien.query.filter_by(masothe=masothe, macongty=macongty).first()    
            if user and user.matkhau == matkhau:
                if login_user(user):    
                    print(f"Nguoi dung {current_user.masothe} o {current_user.macongty} vua  dang nhap !!!")
                    return redirect(url_for('home'))
            return redirect(url_for("login"))
        except Exception as e:
            print(f'Nguoi dung {masothe} o {macongty} dang nhap that bai: {e} !!!')
            return redirect(url_for("login"))
    else:
        danhsachcongty = ["NT1","NT2"]
        return render_template("login.html", danhsachcongty=danhsachcongty)

@app.route("/logout", methods=["GET","POST"])
@login_required
def logout():
    try:
        print(f"Nguoi dung {current_user.masothe} o {current_user.macongty} vua  dang xuat !!!")
        logout_user()
    except Exception as e:
        print(f'Không thế đăng xuất {e} !!!')
    return redirect("/")

@app.route("/", methods=['GET','POST'])
@login_required
def home():
    if request.method == "GET":
        ngay = request.args.get("ngay")   
        chuyen = request.args.get('chuyen')
        style = request.args.get("style")
        mst = request.args.get("mst")
        hoten = request.args.get("hoten")
        macongdoan = request.args.get("search_macongdoan")
        styles = get_all_styles(ngay, chuyen)
        sanluongtong = lay_sanluong_tong_theochuyen(ngay, chuyen, style)
        if g.notice['role'] == 'tt':
            chuyen = g.notice['line'][0]
        danhsach_congnhan_hotro = lay_danhsach_congnhan_trongchuyen(chuyen)
        danhsach_chuyen = lay_danhsach_chuyen_hotro(chuyen)
        danhsach_sanluong = lay_danhsach_sanluong(ngay, chuyen, style,mst,hoten,macongdoan)
        danhsach_tnc = lay_danhsach_tnc_chua_lenchuyen(chuyen)
        danhsach_tnc_ngoichuyen = lay_danhsach_tnc_ngoichuyen(chuyen)
        danhsach_di_hotro = lay_danhsach_di_hotro(chuyen)
        danhsach_den_hotro = lay_danhsach_den_hotro(chuyen)
        return render_template("home.html",styles=styles,danhsach_sanluong=danhsach_sanluong,
                               danhsach_congnhan_hotro=danhsach_congnhan_hotro,
                               danhsach_chuyen=danhsach_chuyen,danhsach_tnc=danhsach_tnc,
                               danhsach_di_hotro=danhsach_di_hotro,sanluongtong=sanluongtong,
                               danhsach_tnc_ngoichuyen=danhsach_tnc_ngoichuyen,
                               danhsach_den_hotro=danhsach_den_hotro)
    elif request.method == "POST":
        try:
            ngay = request.form.get("ngay")   
            chuyen = request.form.get('chuyen')
            style = request.form.get("style")
            mst = request.form.get("mst")
            hoten = request.form.get("hoten")
            macongdoan = request.form.get("search_macongdoan")
            danhsach_sanluong = lay_danhsach_sanluong(ngay, chuyen, style,mst,hoten,macongdoan)
            data = [{
                "Mã số thẻ": int(row[0]),
                "Họ tên": row[1],
                "Chuyền": row[2],
                "Ngày": datetime.datetime.strptime(row[3], "%Y-%m-%d").strftime("%d/%m/%Y"),
                "Style": row[4],
                "Mã công đoạn": int(row[5]) if row[5] else 0,
                "Sản lượng": int(row[6]) if row[6] else 0,
            } for row in danhsach_sanluong]
            df = DataFrame(data)
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename={ngay}_{chuyen}_{style}_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response    
        except Exception as e:
            print(f'Không thế tạo bảng {e} !!!')
            return redirect("/")    
    
@app.route("/nhapsanluongcanhan", methods=["POST"])
@login_required
def nhapsanluongcanhan():
    if request.method == "POST":
        ngay = request.form.get("ngay")   
        chuyen = request.form.get("chuyen")
        style = request.form.get("style")
        mst = request.form.get("mst")
        hoten = request.form.get("hoten")
        macongdoan = request.form.get("macongdoan")
        sanluong = request.form.get("sanluong")
        capnhat_sanluong(mst,hoten,chuyen,ngay,style,macongdoan,sanluong)   
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}&search_macongdoan={macongdoan}")
    
@app.route("/nhapsanluong_congdoanmoi", methods=["POST"])
@login_required
def nhapsanluong_congdoanmoi():
    if request.method == "POST":
        ngay = request.form.get("ngay")   
        chuyen = request.form.get("chuyen")
        style = request.form.get("style")
        mst = request.form.get("mst")
        hoten = request.form.get("hoten")
        macongdoan = request.form.get("macongdoan")
        sanluong = request.form.get("sanluong")
        capnhat_sanluong(mst,hoten,chuyen,ngay,style,macongdoan,sanluong)   
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}&mst={mst}")

@app.route('/xemtencongdoan', methods=["GET","POST"])
@login_required
def xemtencongdoan():
    thongtin = request.args.get("thongtin")
    tencongdoan = lay_tencongdoan(thongtin)
    if tencongdoan:
        return jsonify(tencongdoan)
    else:
        return jsonify("Không tìm thấy")
    
@app.route("/themnguoidihotro", methods=["POST"])
@login_required
def themnguoidihotro():
    if request.method == "POST":
        chuyen = request.form.get("line_dieuchuyendi")
        mst = request.form.get("nguoiduocdieuchuyen").split("_")[0]
        hoten = request.form.get("nguoiduocdieuchuyen").split("_")[1]
        chucdanh = request.form.get("nguoiduocdieuchuyen").split("_")[2]
        chuyendihotro = request.form.get("chuyenhotro")
        ngaydieuchuyendi = request.form.get("ngaydieuchuyendi")
        giodieuchuyendi = request.form.get("giodieuchuyendi")
        sogiohotro = request.form.get("sogiohotro").replace(",", ".")
        nhamay = 'NT1' if chuyen[0]=="1" else 'NT2'
        them_nguoi_di_hotro(nhamay,chuyen,mst,hoten,chucdanh,chuyendihotro,ngaydieuchuyendi,giodieuchuyendi,sogiohotro)
        ngay = request.form.get("ngay")   
        chuyen = request.args.get('chuyen')
        style = request.form.get("style")
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}")

@app.route("/nhantnclenchuyen", methods=["POST"])
@login_required
def nhantnclenchuyen():
    if request.method == "POST":
        id = request.form.get("id_tnc")
        chuyen = request.form.get("chuyen_nhan_tnc")
        nhan_tnc_len_chuyen(id,chuyen)
        ngay = request.form.get("ngay")   
        chuyen = request.args.get('chuyen')
        style = request.form.get("style")
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}")

@app.route("/laytongsanluongtheocongdoan", methods=["POST"])
@login_required
def laytongsanluong():
    if request.method == "POST":
        ngay = request.args.get("ngay")
        chuyen = request.args.get("chuyen")
        style = request.args.get("style")
        macongdoan = request.args.get("macongdoan")
        data = laytongsanluongtheocongdoan(ngay,chuyen,style,macongdoan)
        return jsonify(data)
    
@app.route("/capnhatsogiohotro", methods=["POST"])
@login_required
def capnhatsogiohotro():
    if request.method == "POST":
        id = request.form.get("id_hotro")
        sogio = request.form.get("sogio")
        capnhat_sogio_hotro(id,sogio)
        ngay = request.form.get("ngay")   
        chuyen = request.args.get('chuyen')
        style = request.form.get("style")
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}")
        
@app.route("/capnhatsogiodenhotro", methods=["POST"])
@login_required
def capnhatsogiodenhotro():
    if request.method == "POST":
        id = request.form.get("id_hotro")
        sogio = request.form.get("sogio")
        capnhat_sogio_den_hotro(id,sogio)
        ngay = request.form.get("ngay")   
        chuyen = request.args.get('chuyen')
        style = request.form.get("style")
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}")

@app.route("/xoasanluongcanhan", methods=["POST"])
@login_required
def xoasanluongcanhan():
    if request.method == "POST":
        id = request.form.get("id_xoasanluong")
        ngay = request.form.get("ngay")   
        chuyen = request.form.get('chuyen')
        style = request.form.get("style")
        xoa_sanluong(id)
        return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}")
        
@app.route("/taidulieulen", methods=["POST"])
def taidulieulen():
    if request.method == "POST":
        try:
            file = request.files["file"]
            thoigian = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            filepath = f"tailen/data_{thoigian}.xlsx"
            file.save(filepath)
            data = read_excel(filepath).to_dict(orient="records")
            x=1
            for row in data:
                # print(row['Ngày'])
                ketqua =  capnhat_sanluong(
                    row["Mã số thẻ"],
                    row["Họ tên"],
                    row["Chuyền"],
                    datetime.datetime.strptime(row["Ngày"],"%d/%m/%Y").strftime("%Y-%m-%d") if type(row['Ngày'])==str else row['Ngày'].date(),
                    row["Style"],
                    row["Mã công đoạn"],
                    row["Sản lượng"])
                # print(ketqua)
                if not ketqua["ketqua"]:
                    print(f"Loi dong so {x}, li do {ketqua["lido"]}, query {ketqua["query"]}")
        except Exception as e:
            print(e)
    chuyen = request.form.get('chuyen')
    ngay = request.form.get('ngay')
    style = request.form.get('style')
    return redirect(f"/?chuyen={chuyen}&ngay={ngay}&style={style}")

@app.route("/baocao_thuong_cat", methods=["GET","POST"])
def baocao_cat():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_cat(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_cat.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_cat.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_cat(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Chuyền":row[2],
                    "Ngày":datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Hệ số":round(row[4],2) if row[4] else "",
                    "Số giờ":row[5],
                    "Hệ số cá nhân": round(row[6],2) if row[6] else "",
                    "Tổng hệ số": round(row[7],2) if row[7] else "",
                    "Thưởng nhóm":round(row[8]) if row[8] else "",
                    "Thưởng cá nhân": round(row[9]) if row[9] else ""
                })
            df = DataFrame(data)
            # Sử dụng pd.to_numeric để chuyển đổi cột 'A' sang int64
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Hệ số'] = to_numeric(df['Hệ số'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hệ số cá nhân'] = to_numeric(df['Hệ số cá nhân'], errors='coerce')
            df['Tổng hệ số'] = to_numeric(df['Tổng hệ số'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongcat_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_cat")

@app.route("/baocao_thuong_may", methods=["GET","POST"])
def baocao_may():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_may(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_may.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_may.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_may(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Chuyền":row[2],
                    "Ngày":datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "SAH":round(row[4],2) if row[4] else "",
                    "SCP":row[5],
                    "Số giờ":row[6],
                    "Hiệu suất": round(row[7],2) if row[7] else "",
                    "Thưởng": round(row[8]) if row[8] else ""
                })
            df = DataFrame(data)
            # Sử dụng pd.to_numeric để chuyển đổi cột 'A' sang int64
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Thưởng'] = to_numeric(df['Thưởng'], errors='coerce')
            df['Hiệu suất'] = to_numeric(df['Hiệu suất'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongmay_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_may")

@app.route("/baocao_thuong_la", methods=["GET","POST"])
def baocao_la():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_la(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_la.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_la.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_la(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Ngày": datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[2],
                    "SAH": round(row[4],2) if row[4] else "",
                    "Số giờ":row[5],
                    "Hiệu suất cá nhân": round(row[6],2) if row[6] else "",
                    "SAH nhóm": round(row[7],2) if row[7] else "",
                    "Thưởng nhóm": round(row[8]) if row[8] else "",
                    "Thưởng cá nhân": round(row[9]) if row[9] else ""
                })
            df = DataFrame(data)
            # Sử dụng pd.to_numeric để chuyển đổi cột 'A' sang int64
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['SAH nhóm'] = to_numeric(df['SAH nhóm'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Hiệu suất cá nhân'] = to_numeric(df['Hiệu suất cá nhân'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongla_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_la")
       
@app.route("/baocao_thuong_qc1", methods=["GET","POST"])
def baocao_qc1():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_qc1(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_qc1.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_qc1.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_qc1(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Ngày": datetime.datetime.strptime(row[2],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[3],
                    "SAH": round(row[4],2) if row[4] else "",
                    "Số giờ":row[5],
                    "Hiệu suất cá nhân": round(row[6],2) if row[6] else "",
                    "SAH nhóm": round(row[7],2) if row[7] else "",
                    "Thưởng nhóm": round(row[8]) if row[8] else "",
                    "Thưởng cá nhân": round(row[9]) if row[9] else ""
                })
            df = DataFrame(data)
            # Sử dụng pd.to_numeric để chuyển đổi cột 'A' sang int64
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['SAH nhóm'] = to_numeric(df['SAH nhóm'], errors='coerce')
            df['Hiệu suất cá nhân'] = to_numeric(df['Hiệu suất cá nhân'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongqc1_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_qc1")
        
@app.route("/baocao_thuong_qc2", methods=["GET","POST"])
def baocao_qc2():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_qc2(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_qc2.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_qc2.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_qc2(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Ngày": datetime.datetime.strptime(row[2],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[3],
                    "SAH": round(row[4],2) if row[4] else "",
                    "Số giờ":row[5],
                    "Hiệu suất cá nhân": round(row[6],2) if row[6] else "",
                    "SAH nhóm": round(row[7],2) if row[7] else "",
                    "Thưởng nhóm": round(row[8]) if row[8] else "",
                    "Thưởng cá nhân": round(row[9]) if row[9] else ""
                })
            df = DataFrame(data)
            # Sử dụng pd.to_numeric để chuyển đổi cột 'A' sang int64
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['SAH nhóm'] = to_numeric(df['SAH nhóm'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hiệu suất cá nhân'] = to_numeric(df['Hiệu suất cá nhân'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongqc1_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_qc2")
            
@app.route("/baocao_thuong_donggoi", methods=["GET","POST"])
def baocao_donggoi():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_donggoi(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_donggoi.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_donggoi.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_donggoi(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Ngày": datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[2],
                    "Số giờ":row[4],
                    "OT": row[6],
                    "Hệ số cá nhân": round(row[7],2) if row[7] else "",
                    "Hệ số nhóm": round(row[9],2) if row[3] else "",
                    "Thưởng nhóm": round(row[8]) if row[8] else "",
                    "Thưởng cá nhân": round(row[10]) if row[10] else ""
                })
            df = DataFrame(data)
            # Sử dụng pd.to_numeric để chuyển đổi cột 'A' sang int64
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['OT'] = to_numeric(df['OT'], errors='coerce')
            df['Hệ số cá nhân'] = to_numeric(df['Hệ số cá nhân'], errors='coerce')
            df['Hệ số nhóm'] = to_numeric(df['Hệ số nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongdonggoi_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_donggoi")

@app.route("/baocao_thuong_ndc", methods=["GET","POST"])
def baocao_ndc():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_ndc(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_ndc.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_ndc.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_ndc(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Ngày": datetime.datetime.strptime(row[4],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[3],
                    "Hệ số cá nhân": round(row[5],2) if row[5] else "",
                    "Hệ số nhóm": round(row[7],2) if row[7] else "",
                    "Thưởng nhóm": round(row[6]) if row[6] else "",
                    "Thưởng cá nhân": round(row[8]) if row[8] else ""
                })
            df = DataFrame(data)
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Hệ số cá nhân'] = to_numeric(df['Hệ số cá nhân'], errors='coerce')
            df['Hệ số nhóm'] = to_numeric(df['Hệ số nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongndc_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_ndc")
        
@app.route("/baocao_thuong_cnphu", methods=["GET","POST"])
def baocao_cn_phu():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_phu(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_cnphu.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_cnphu.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_phu(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[2],
                    "Họ tên":row[3],
                    "Ngày": row[0],
                    "Chuyền":row[1],
                    "Chức danh": row[4],
                    "Số giờ": round(row[5],2) if row[5] else 0.0,
                    "Hệ số": round(row[6],2) if row[6] else 0.0,
                    "Đánh giá": row[7] if row[7] else "",
                    "Hệ số đánh giá": round(row[8],2) if row[8] else 0.0,
                    "Tổng TGLV nhóm": round(row[9]) if row[9] else 0.0,
                    "Thưởng nhóm": round(row[10]) if row[10] else 0,
                    "Thưởng cá nhân": round(row[11]) if row[11] else 0
                })
            df = DataFrame(data)
            df['Ngày'] = to_datetime(df['Ngày'], errors='coerce')
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hệ số'] = to_numeric(df['Hệ số'], errors='coerce')
            df['Hệ số đánh giá'] = to_numeric(df['Hệ số đánh giá'], errors='coerce')
            df['Hệ số đánh giá'] = to_numeric(df['Hệ số đánh giá'], errors='coerce')
            df['Tổng TGLV nhóm'] = to_numeric(df['Tổng TGLV nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongcnphu_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_cnphu")
        
@app.route("/baocao_thuong_quanly", methods=["GET","POST"])
def baocao_quanly():
    if request.method == "GET":
        if ((current_user.phanquyen=="sa" or current_user.phanquyen=="gd") or ("IE" in current_user.phongban and not "W" in current_user.capbac)):
            try:
                macongty = request.args.get("macongty")
                mst = request.args.get("mst")
                ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
                chuyen = request.args.get("chuyen")
                danhsach = lay_baocao_thuong_quanly(macongty,mst,ngay,chuyen)
                page = request.args.get(get_page_parameter(), type=int, default=1)
                per_page = 10
                total = len(danhsach)
                start = (page - 1) * per_page
                end = start + per_page
                paginated_rows = danhsach[start:end]
                pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
                return render_template("baocao_thuong_quanly.html", danhsach=paginated_rows,pagination=pagination)
            except Exception as e:
                print(e)
                return render_template("baocao_thuong_quanly.html", danhsach=[])
        else:
            return redirect("/baocao_thuong_may")
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_quanly(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[2],
                    "Họ tên":row[3],
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "Chức vụ": row[4],
                    "Số giờ": row[5],
                    "Hệ số": row[6],
                    "Đánh giá": row[7], 
                    "Hệ số đánh giá": row[8] if row[8] else "",
                    "Tổng TGLV nhóm": row[9] if row[9] else "",
                    "Thưởng nhóm": round(row[10]) if row[10] else "",
                    "Thưởng cá nhân": round(row[11]) if row[11] else ""
                })
            df = DataFrame(data)
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Hệ số'] = to_numeric(df['Hệ số'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hệ số đánh giá'] = to_numeric(df['Hệ số đánh giá'], errors='coerce')
            df['Tổng TGLV nhóm'] = to_numeric(df['Tổng TGLV nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Thưởng cá nhân'] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongquanly_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_quanly")
       
@app.route("/baocao_thuong_nhommay", methods=["GET","POST"])
def baocao_nhommay():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            style = request.args.get("style")
            danhsach = lay_baocao_thuong_congnhan_nhommay(macongty,ngay,chuyen,style)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhommay.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_may.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            style = request.form.get("style")
            danhsach = lay_baocao_thuong_congnhan_nhommay(macongty,ngay,chuyen,style)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "SAH": round(row[2],2) if row[2] else "",
                    "Số giờ":row[3],
                    "Hiệu suất": round(row[4],2) if row[4] else "",
                    "Style": row[5],
                    "Trạng thái đơn hàng": row[6],
                    "Chuyền mới": row[7],
                    "OQL": row[8],
                    "Thưởng nhóm":round(row[9]) if row[9] else "",
                    "Thưởng 1": round(row[10]) if row[10] else "",
                    "Thưởng 2": round(row[11]) if row[11] else "",
                    "Tổng thưởng": round(row[12]) if row[12] else ""
                })
            df = DataFrame(data)
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hiệu suất'] = to_numeric(df['Hiệu suất'], errors='coerce')
            df['Thưởng 1'] = to_numeric(df['Thưởng 1'], errors='coerce')
            df['Thưởng 2'] = to_numeric(df['Thưởng 2'], errors='coerce')
            df['Tổng thưởng'] = to_numeric(df['Tổng thưởng'], errors='coerce')
            output = BytesIO()
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhommay_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhommay")

@app.route("/baocao_thuong_nhomcat", methods=["GET","POST"])
def baocao_nhomcat():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomcat(macongty,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhomcat.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_nhomcat.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomcat(macongty,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "SAH":round(row[2],2) if row[2] else "",
                    "Số giờ":round(row[3],1),
                    "Hiệu suất": round(row[4],2) if row[4] else "",
                    "Cắt kẻ": row[5],
                    "Nhà máy": row[6],
                    "UI": row[7],
                    "BE": round(row[8],2) if row[8] else "",
                    "BE TOPUP 1":round(row[9],2) if row[9] else "",
                    "TOP UP 1": round(row[10],2) if row[10] else "",
                    "Tỉ lệ lỗi": round(row[11],2) if row[11] else "",
                    "AQL": round(row[12],2) if row[12] else "",
                    "Thưởng nhóm": round(row[13]) if row[13] else "",
                    "Kích cầu 1": round(row[14]) if row[14] else "",
                    "Tổng thưởng nhóm": round(row[15]) if row[15] else "",
                })
            df = DataFrame(data)
            
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['UI'] = to_numeric(df['UI'], errors='coerce')
            df['Kích cầu 1'] = to_numeric(df['Kích cầu 1'], errors='coerce')
            df['Tổng thưởng nhóm'] = to_numeric(df['Tổng thưởng nhóm'], errors='coerce')
            df['Hiệu suất'] = to_numeric(df['Hiệu suất'], errors='coerce')
            df['BE'] = to_numeric(df['BE'], errors='coerce')
            df['BE TOPUP 1'] = to_numeric(df['BE TOPUP 1'], errors='coerce')
            df['TOP UP 1'] = to_numeric(df['TOP UP 1'], errors='coerce')
            df['Tỉ lệ lỗi'] = to_numeric(df['Tỉ lệ lỗi'], errors='coerce')
            df['AQL'] = to_numeric(df['AQL'], errors='coerce')
            output = BytesIO()
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhomcat_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhomcat")

@app.route("/baocao_thuong_nhomla", methods=["GET","POST"])
def baocao_nhomla():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomla(macongty,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhomla.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_nhomla.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomla(macongty,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "SAH":round(row[2],2) if row[2] else "",
                    "Số giờ":round(row[3],1),
                    "Hiệu suất": round(row[4],2)if row[4] else "",
                    "Nhà máy": row[5],
                    "UI": row[6],
                    "BE": round(row[7],2) if row[7] else "",
                    "BE TOPUP 1":round(row[8],2) if row[8] else "",
                    "TOP UP 1": round(row[9],2) if row[9] else "",
                    "BE TOPUP 2":round(row[10],2) if row[10] else "",
                    "TOP UP 2": round(row[11],2) if row[11] else "",
                    "Tỉ lệ lỗi": round(row[12],2) if row[12] else "",
                    "AQL": round(row[13],2) if row[13] else "",
                    "Thưởng nhóm": round(row[14]) if row[14] else "",
                    "Kích cầu 1": round(row[15]) if row[15] else "",
                    "Kích cầu 2": round(row[16]) if row[16] else "",
                    "Tổng thưởng nhóm": round(row[17]) if row[17] else "",
                })
            df = DataFrame(data)
            
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['UI'] = to_numeric(df['UI'], errors='coerce')
            df['Kích cầu 1'] = to_numeric(df['Kích cầu 1'], errors='coerce')
            df['Kích cầu 2'] = to_numeric(df['Kích cầu 2'], errors='coerce')
            df['Tổng thưởng nhóm'] = to_numeric(df['Tổng thưởng nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['Hiệu suất'] = to_numeric(df['Hiệu suất'], errors='coerce')
            df['BE'] = to_numeric(df['BE'], errors='coerce')
            df['BE TOPUP 1'] = to_numeric(df['BE TOPUP 1'], errors='coerce')
            df['TOP UP 1'] = to_numeric(df['TOP UP 1'], errors='coerce')
            df['BE TOPUP 2'] = to_numeric(df['BE TOPUP 2'], errors='coerce')
            df['TOP UP 2'] = to_numeric(df['TOP UP 2'], errors='coerce')
            df['Tỉ lệ lỗi'] = to_numeric(df['Tỉ lệ lỗi'], errors='coerce')
            df['AQL'] = to_numeric(df['AQL'], errors='coerce')
            output = BytesIO()
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhomla_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhomla")

@app.route("/baocao_thuong_nhomdonggoi", methods=["GET","POST"])
def baocao_nhomdonggoi():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomdonggoi(macongty,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhomdonggoi.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_nhomdonggoi.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomdonggoi(macongty,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "SAH":round(row[2],2) if row[2] else "",
                    "Số giờ":round(row[3],1),
                    "Hiệu suất": round(row[4],2) if row[4] else "",
                    "Nhà máy": row[5],
                    "UI": row[6],
                    "BE": f"{row[7]:.0%}" if row[7] else "",
                    "BE TOPUP 1":round(row[8],2)if row[8] else "",
                    "TOP UP 1": round(row[9],2) if row[9] else "",
                    "Tỉ lệ lỗi": round(row[10],2) if row[10] else "",
                    "AQL": round(row[11],2) if row[11] else "",
                    "Thưởng nhóm": round(row[12]) if row[12] else "",
                    "Kích cầu 1": round(row[13]) if row[13] else "",
                    "Tổng thưởng nhóm": round(row[14]) if row[14] else "",
                })
            df = DataFrame(data)
            
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['UI'] = to_numeric(df['UI'], errors='coerce')
            df['Kích cầu 1'] = to_numeric(df['Kích cầu 1'], errors='coerce')
            df['Tổng thưởng nhóm'] = to_numeric(df['Tổng thưởng nhóm'], errors='coerce')
            output = BytesIO()
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhomdonggoi_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhomdonggoi")

@app.route("/baocao_thuong_nhomndc", methods=["GET","POST"])
def baocao_nhomndc():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomndc(macongty,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhomndc.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_nhomndc.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomndc(macongty,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": row[1],
                    "Chuyền":row[0],
                    "Số công nhân":row[2],
                    "Nhà máy": row[3],
                    "UI": row[4],
                    "OT FNS": row[5],
                    "OT NDC": row[6],
                    "Hệ số OT": row[7],
                    "Số ngày công": row[8],
                    "Tổng thưởng nhóm": round(row[9]) if row[9] else 0,
                })
            df = DataFrame(data)
            
            df['Tổng thưởng nhóm'] = to_numeric(df['Tổng thưởng nhóm'], errors='coerce')
            df['UI'] = to_numeric(df['UI'], errors='coerce')
            df['OT FNS'] = to_numeric(df['OT FNS'], errors='coerce')
            df['OT NDC'] = to_numeric(df['OT NDC'], errors='coerce')
            df['Hệ số OT'] = to_numeric(df['Hệ số OT'], errors='coerce')
            df['Số ngày công'] = to_numeric(df['Số ngày công'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhomndc_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhomndc")
        
@app.route("/baocao_thuong_nhomqc1", methods=["GET","POST"])
def baocao_nhomqc1():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomqc1(macongty,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhomqc1.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_nhomqc1.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomqc1(macongty,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "SAH":round(row[2],2) if row[2] else "",
                    "Số giờ":round(row[3],1),
                    "Hiệu suất": round(row[4],2) if row[4] else "",
                    "Nhà máy": row[5],
                    "UI": row[6],
                    "BE": round(row[7],2) if row[7] else "",
                    "BE TOPUP 1":round(row[8],2) if row[8] else "",
                    "TOP UP 1": round(row[9],2) if row[9] else "",
                    "Tỉ lệ lỗi": round(row[10],2) if row[10] else "",
                    "AQL": round(row[11],2) if row[11] else "",
                    "Thưởng nhóm": round(row[12]) if row[12] else "",
                    "Kích cầu 1": round(row[13]) if row[13] else "",
                    "Tổng thưởng nhóm": round(row[14]) if row[14] else "",
                })
            df = DataFrame(data)
            
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hiệu suất'] = to_numeric(df['Hiệu suất'], errors='coerce')
            df['UI'] = to_numeric(df['UI'], errors='coerce')
            df['Kích cầu 1'] = to_numeric(df['Kích cầu 1'], errors='coerce')
            df['Tổng thưởng nhóm'] = to_numeric(df['Tổng thưởng nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['BE'] = to_numeric(df['BE'], errors='coerce')
            df['BE TOPUP 1'] = to_numeric(df['BE TOPUP 1'], errors='coerce')
            df['TOP UP 1'] = to_numeric(df['TOP UP 1'], errors='coerce')
            df['Tỉ lệ lỗi'] = to_numeric(df['Tỉ lệ lỗi'], errors='coerce')
            df['AQL'] = to_numeric(df['AQL'], errors='coerce')
            output = BytesIO()
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhomqc1_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhomqc1")
        
@app.route("/baocao_thuong_nhomqc2", methods=["GET","POST"])
def baocao_nhomqc2():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomqc2(macongty,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_thuong_nhomqc2.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_thuong_nhomcqc2.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_thuong_congnhan_nhomqc2(macongty,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Chuyền":row[1],
                    "SAH":round(row[2],2) if row[2] else "",
                    "Số giờ":round(row[3],1),
                    "Hiệu suất": round(row[4],2) if row[4] else "",
                    "Nhà máy": row[5],
                    "UI": row[6],
                    "BE": round(row[7],2) if row[7] else "",
                    "BE TOPUP 1":round(row[8],2) if row[8] else "",
                    "TOP UP 1": round(row[9],2) if row[9] else "",
                    "Tỉ lệ lỗi": round(row[10],2) if row[10] else "",
                    "AQL": round(row[11],2) if row[11] else "",
                    "Thưởng nhóm": round(row[12]) if row[12] else "",
                    "Kích cầu 1": round(row[13]) if row[13] else "",
                    "Tổng thưởng nhóm": round(row[14]) if row[14] else "",
                })
            df = DataFrame(data)
            
            df['SAH'] = to_numeric(df['SAH'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            df['Hiệu suất'] = to_numeric(df['Hiệu suất'], errors='coerce')
            df['UI'] = to_numeric(df['UI'], errors='coerce')
            df['Kích cầu 1'] = to_numeric(df['Kích cầu 1'], errors='coerce')
            df['Tổng thưởng nhóm'] = to_numeric(df['Tổng thưởng nhóm'], errors='coerce')
            df['Thưởng nhóm'] = to_numeric(df['Thưởng nhóm'], errors='coerce')
            df['BE'] = to_numeric(df['BE'], errors='coerce')
            df['BE TOPUP 1'] = to_numeric(df['BE TOPUP 1'], errors='coerce')
            df['TOP UP 1'] = to_numeric(df['TOP UP 1'], errors='coerce')
            df['Tỉ lệ lỗi'] = to_numeric(df['Tỉ lệ lỗi'], errors='coerce')
            df['AQL'] = to_numeric(df['AQL'], errors='coerce')
            output = BytesIO()
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaothuongnhomqc2_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_thuong_nhomqc2")

@app.route("/baocao_sogio_lamviec", methods=["GET", "POST"])
def baocao_sogio_lamviec():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_sogio_lamviec(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_sogio_lamviec.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_sogio_lamviec.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_sogio_lamviec(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên":row[1],
                    "Chuyền":row[2],
                    "Ngày": datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Số giờ":row[4],
                    "Chức danh" : row[5],
                })
            df = DataFrame(data)
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Số giờ'] = to_numeric(df['Số giờ'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaosogiolamviec_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_sogio_lamviec")
    
@app.route("/baocao_sanluong_canhan", methods=["GET","POST"])
def baocao_sanluong_canhan():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            ngay = request.args.get("ngay") if request.args.get("ngay") else datetime.datetime.now().strftime("%Y-%m-%d")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_sanluong_canhan(macongty,mst,ngay,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_sanluong_canhan.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_sanluong_canhan.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            ngay = request.form.get("ngay")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_sanluong_canhan(macongty,mst,ngay,chuyen)
            data = []
            for row in danhsach:
                data.append({
                    "Mã số thẻ": row[0],
                    "Họ tên": row[1],
                    "Chuyền":row[2],
                    "Ngày": datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Style":row[4],
                    "Mã công đoạn" : row[5],
                    "Sản lượng": row[6]
                })
            df = DataFrame(data)
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            df['Mã công đoạn'] = to_numeric(df['Mã công đoạn'], errors='coerce')
            df['Sản lượng'] = to_numeric(df['Sản lượng'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaosanluongcanhan_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_sanluong_canhan")

@app.route("/baocao_scp_canhan", methods=["GET","POST"])
def baocao_scp_canhan():
    if request.method == "GET":
        try:
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            tungay = request.args.get("tungay")
            denngay = request.args.get("denngay") 
            danhsach = lay_baocao_scp_canhan(macongty,mst,tungay,denngay)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_scp_canhan.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_scp_canhan.html", danhsach=[])
    elif request.method == "POST":
        try:
            macongty = request.form.get("macongty")
            mst = request.args.form("mst")
            tungay = request.form.get("tungay")
            denngay = request.form.get("denngay") 
            danhsach = lay_baocao_scp_canhan(macongty,mst,tungay,denngay)
            data = []
            for row in danhsach:
                data.append({
                    "Nhà máy": row[0],
                    "Mã số thẻ": row[1],
                    "Họ tên": row[2],
                    "Từ ngày": datetime.datetime.strptime(row[3],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Đến ngày": datetime.datetime.strptime(row[4],"%Y-%m-%d").strftime("%d/%m/%Y"),
                    "SCP" : row[5]
                })
            df = DataFrame(data)
            df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Điều chỉnh độ rộng cột
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocaoscpcanhan_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_scp_canhan")
       
@app.route("/taithuongchitiet", methods=["GET","POST"])
def taithuongchitiet():
    try:
        danhsach = laydulieuthuongmaychitiet()
        data = []
        for row in danhsach:
            data.append({
                "Ngày": datetime.datetime.strptime(row[0],"%Y-%m-%d").strftime("%d/%m/%Y"),
                "Mã số thẻ": row[1],
                "Họ tên": row[2],
                "Chuyền": row[3],
                "SAH nhóm": round(row[4],1) if row[4] else "",
                "Thời gian làm việc nhóm": round(row[5]) if row[5] else "",
                "Hiệu suất nhóm": round(row[6],2) if row[6] else "",
                "Style": row[7] if row[7] else "",
                "Chuyền mới": row[8] if row[8] else "",
                "Ngày vào chuyền": row[9] if row[9] else "",
                "Trạng thái": row[10] if row[10] else "",
                "UI": row[11] if row[11] else "",
                "BE": round(row[12],2) if row[12] else "",
                "BE TOPUP 1": round(row[13],2) if row[13] else "",
                "TOPUP 1": round(row[14],2) if row[14] else "",
                "BE TOPUP 2": round(row[15],2) if row[15] else "",
                "TOPUP 2": round(row[16],2) if row[16] else "",
                "OQL": round(row[17],2) if row[17] else "",
                "AQL": round(row[18],2) if row[18] else "",
                "Group incentive": round(row[19]) if row[19] else "",
                "Group incentive topup 1": round(row[20]) if row[20] else "",
                "Group incentive topup 2": round(row[21]) if row[21] else "",
                "Tổng thưởng": round(row[22]) if row[22] else "",
                "SAH": round(row[23],1) if row[23] else "",
                "Thời gian làm việc": row[24] if row[24] else "",
                "Hiệu suất": round(row[25],2) if row[25] else "",
                "SCP": row[26] if row[26] else "",
                "Hệ số SCP": row[27] if row[27] else "",
                "Hệ số thưởng cá nhân": round(row[28],1) if row[28] else "",
                "Hệ số thưởng nhóm": round(row[29],1) if row[29] else "",
                "Thưởng cá nhân": round(row[30]) if row[30] else ""
            })
        df = DataFrame(data)
        df["Mã số thẻ"] = to_numeric(df['Mã số thẻ'], errors='coerce')
        df["SAH nhóm"] = to_numeric(df['SAH nhóm'], errors='coerce')
        df["Thời gian làm việc nhóm"] = to_numeric(df['Thời gian làm việc nhóm'], errors='coerce')
        df["Group incentive"] = to_numeric(df['Group incentive'], errors='coerce')
        df["Group incentive topup 1"] = to_numeric(df['Group incentive topup 1'], errors='coerce')
        df["Group incentive topup 2"] = to_numeric(df['Group incentive topup 2'], errors='coerce')
        df["Tổng thưởng"] = to_numeric(df['Tổng thưởng'], errors='coerce')
        df["SAH"] = to_numeric(df['SAH'], errors='coerce')
        df["Thời gian làm việc"] = to_numeric(df['Thời gian làm việc'], errors='coerce')
        df["Hệ số SCP"] = to_numeric(df['Hệ số SCP'], errors='coerce')
        df["Hệ số thưởng cá nhân"] = to_numeric(df['Hệ số thưởng cá nhân'], errors='coerce')
        df["Hệ số thưởng nhóm"] = to_numeric(df['Hệ số thưởng nhóm'], errors='coerce')
        df["Thưởng cá nhân"] = to_numeric(df['Thưởng cá nhân'], errors='coerce')
        df["Hiệu suất nhóm"] = to_numeric(df['Hiệu suất nhóm'], errors='coerce')
        df["BE"] = to_numeric(df['BE'], errors='coerce')
        df["BE TOPUP 1"] = to_numeric(df['BE TOPUP 1'], errors='coerce')
        df["TOPUP 1"] = to_numeric(df['TOPUP 1'], errors='coerce')
        df["BE TOPUP 2"] = to_numeric(df['BE TOPUP 2'], errors='coerce')
        df["TOPUP 2"] = to_numeric(df['TOPUP 2'], errors='coerce')
        df["OQL"] = to_numeric(df['OQL'], errors='coerce')
        df["AQL"] = to_numeric(df['AQL'], errors='coerce')
        df["Hiệu suất"] = to_numeric(df['Hiệu suất'], errors='coerce')
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        # Điều chỉnh độ rộng cột
        output.seek(0)
        workbook = load_workbook(output)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
        # Trả file về cho client
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=thuongmaychitiet_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response  
    except Exception as e:
        print(f"Loi lay thuong may chi tiet: {e}")
        return redirect("/")

@app.route("/capnhat_tungay_tnc", methods=["POST"])
def capnhat_tungay_cn_tnc():
    id = request.form.get("id")
    ngay = request.form.get("ngay")
    try:
        thaydoi_tungay_cn_tnc(id, ngay)
    except Exception as e:
        print(f"Loi : {e}")
    return redirect("/")
    
@app.route("/capnhat_denngay_tnc", methods=["POST"])
def capnhat_denngay_cn_tnc():
    id = request.form.get("id")
    ngay = request.form.get("ngay")
    try:
        thaydoi_denngay_cn_tnc(id, ngay)
    except Exception as e:
        print(f"Loi : {e}")
    return redirect("/")
    
@app.route("/danhsach_dihotro", methods=["GET","POST"])
def danhsach_dihotro():
    if request.method == "GET":
        mst = request.args.get("mst")
        chuyen = request.args.get("chuyen")
        ngay = request.args.get("ngay")
        danhsach = lay_danhsach_di_hotro_tatca(mst,chuyen,ngay)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        for row in paginated_rows:
            row_list = list(row)
            row_list[7] = f"{row_list[7].split(":")[0]}:{row_list[7].split(":")[1]}" if row_list[7] else row_list[7]
            paginated_rows[paginated_rows.index(row)] = tuple(row_list)
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("danhsach_dihotro.html", danhsach=paginated_rows,pagination=pagination)
        
    elif request.method == "POST":
        mst = request.form.get("mst")
        chuyen = request.form.get("chuyen")
        ngay = request.form.get("ngay")
        danhsach = lay_danhsach_di_hotro_tatca(mst,chuyen,ngay)
        data = []
        for row in danhsach:
            data.append({
                "Nhà máy": row[0],
                "Mã số thẻ": row[1],
                "Họ tên": row[2],
                "Chức danh": row[3],
                "Chuyền": row[4],
                "Chuyền đi hỗ trợ": row[5],
                "Ngày": row[6],
                "Giờ": row[7],
                "Số giờ": row[8],
                "ID": row[9]
            })
        df = DataFrame(data)
        df["Mã số thẻ"] = to_numeric(df['Mã số thẻ'], errors='coerce')
        df["Số giờ"] = to_numeric(df['Số giờ'], errors='coerce')
        df["Ngày"] = to_datetime(df['Ngày'], errors='coerce').dt.date
        df["Giờ"] = to_datetime(df['Giờ'], errors='coerce').dt.time
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        output.seek(0)
        workbook = load_workbook(output)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=danhsachdihotro_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response  

@app.route("/baocao_hieusuat_may_chitiet", methods=["GET","POST"])
def baocao_hieusuat_may_chitiet():
    if request.method == "GET":
        try:
            nam = request.args.get("nam")
            thang = request.args.get("thang")
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_hieusuat_chitiet_may(nam,thang,macongty,mst,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_hieusuat_may_chitiet.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_hieusuat_may_chitiet.html", danhsach=[])
    elif request.method == "POST":
        try:
            nam = request.form.get("nam")
            thang = request.form.get("thang")
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_hieusuat_chitiet_may(nam,thang,macongty,mst,chuyen)
            data = [{
                "Mã số thẻ": row[3],
                "Họ tên" : row[4],
                "Chuyền" : row[5],
                "01" : row[6],
                "02" : row[7],
                "03" : row[8],
                "04" : row[9],
                "05" : row[10],
                "06" : row[11],
                "07" : row[12],
                "08" : row[13],
                "09" : row[14],
                "10" : row[15],
                "11" : row[16],
                "12" : row[17],
                "13" : row[18],
                "14" : row[19],
                "15" : row[20],
                "16" : row[21],
                "17" : row[22],
                "18" : row[23],
                "19" : row[24],
                "20" : row[25],
                "21" : row[26],
                "22" : row[27],
                "23" : row[28],
                "24" : row[29],
                "25" : row[30],
                "26" : row[31],
                "27" : row[32],
                "28" : row[33],
                "29" : row[34],
                "30" : row[35],
                "31" : row[36],
                "Trung bình" : row[37]
            } for row in danhsach]
            df = DataFrame(data)
            if data:
                df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
                df["01"] = to_numeric(df['01'], errors='coerce') 
                df["02"] = to_numeric(df['02'], errors='coerce') 
                df["03"] = to_numeric(df['03'], errors='coerce') 
                df["04"] = to_numeric(df['04'], errors='coerce') 
                df["05"] = to_numeric(df['05'], errors='coerce') 
                df["06"] = to_numeric(df['06'], errors='coerce') 
                df["07"] = to_numeric(df['07'], errors='coerce') 
                df["08"] = to_numeric(df['08'], errors='coerce')
                df["09"] = to_numeric(df['09'], errors='coerce') 
                df["10"] = to_numeric(df['10'], errors='coerce') 
                df["11"] = to_numeric(df['11'], errors='coerce') 
                df["12"] = to_numeric(df['12'], errors='coerce') 
                df["13"] = to_numeric(df['13'], errors='coerce') 
                df["14"] = to_numeric(df['14'], errors='coerce') 
                df["15"] = to_numeric(df['15'], errors='coerce') 
                df["16"] = to_numeric(df['16'], errors='coerce') 
                df["17"] = to_numeric(df['17'], errors='coerce') 
                df["18"] = to_numeric(df['18'], errors='coerce') 
                df["19"] = to_numeric(df['19'], errors='coerce') 
                df["20"] = to_numeric(df['20'], errors='coerce') 
                df["21"] = to_numeric(df['21'], errors='coerce') 
                df["22"] = to_numeric(df['22'], errors='coerce') 
                df["23"] = to_numeric(df['23'], errors='coerce') 
                df["24"] = to_numeric(df['24'], errors='coerce') 
                df["25"] = to_numeric(df['25'], errors='coerce') 
                df["26"] = to_numeric(df['26'], errors='coerce') 
                df["27"] = to_numeric(df['27'], errors='coerce') 
                df["28"] = to_numeric(df['28'], errors='coerce') 
                df["29"] = to_numeric(df['29'], errors='coerce') 
                df["30"] = to_numeric(df['30'], errors='coerce') 
                df["31"] = to_numeric(df['31'], errors='coerce') 
                df["Trung bình"] = to_numeric(df['Trung bình'], errors='coerce') 
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Adjust column width and format the header row
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            # Style the header row
            header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Create a date format for short date
            date_format = NamedStyle(name="Percentage", number_format="0.00%")
            if "Percentage" not in workbook.named_styles:
                workbook.add_named_style(date_format)
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        # Apply the date format to column L (assuming 'Ngày thực hiện' is in column 'L')
                        if cell.column_letter in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                        'AA','AB','AC','AD','AE','AF','AG','AH','AI'] and cell.value is not None:
                            cell.number_format = '0.00%'
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the modified workbook to the output BytesIO object
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocao_hieusuat_chitiet_may_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_hieusuat_may_chitiet")

@app.route("/baocao_hieusuat_may_tonghop", methods=["GET","POST"])
def baocao_hieusuat_may_tonghop():
    if request.method == "GET":
        try:
            nam = request.args.get("nam")
            thang = request.args.get("thang")
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_hieusuat_tonghop_may(nam,thang,macongty,mst,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_hieusuat_may_tonghop.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_hieusuat_may_tonghop.html", danhsach=[])
    elif request.method == "POST":
        try:
            nam = request.form.get("nam")
            thang = request.form.get("thang")
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_hieusuat_tonghop_may(nam,thang,macongty,mst,chuyen)
            data = [{
                "Mã số thẻ": row[3],
                "Họ tên" : row[4],
                "Chuyền" : row[5],
                "SCP": row[6],
                "01" : row[7],
                "02" : row[8],
                "03" : row[9],
                "04" : row[10],
                "05" : row[11],
                "06" : row[12],
                "07" : row[13],
                "08" : row[14],
                "09" : row[15],
                "10" : row[16],
                "11" : row[17],
                "12" : row[18],
                "13" : row[19],
                "14" : row[20],
                "15" : row[21],
                "16" : row[22],
                "17" : row[23],
                "18" : row[24],
                "19" : row[25],
                "20" : row[26],
                "21" : row[27],
                "22" : row[28],
                "23" : row[29],
                "24" : row[30],
                "25" : row[31],
                "26" : row[32],
                "27" : row[33],
                "28" : row[34],
                "29" : row[35],
                "30" : row[36],
                "31" : row[37],
                "Trung bình" : row[38]
            } for row in danhsach]
            df = DataFrame(data)
            if data:
                df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
                df["01"] = to_numeric(df['01'], errors='coerce') 
                df["02"] = to_numeric(df['02'], errors='coerce') 
                df["03"] = to_numeric(df['03'], errors='coerce') 
                df["04"] = to_numeric(df['04'], errors='coerce') 
                df["05"] = to_numeric(df['05'], errors='coerce') 
                df["06"] = to_numeric(df['06'], errors='coerce') 
                df["07"] = to_numeric(df['07'], errors='coerce') 
                df["08"] = to_numeric(df['08'], errors='coerce')
                df["09"] = to_numeric(df['09'], errors='coerce') 
                df["10"] = to_numeric(df['10'], errors='coerce') 
                df["11"] = to_numeric(df['11'], errors='coerce') 
                df["12"] = to_numeric(df['12'], errors='coerce') 
                df["13"] = to_numeric(df['13'], errors='coerce') 
                df["14"] = to_numeric(df['14'], errors='coerce') 
                df["15"] = to_numeric(df['15'], errors='coerce') 
                df["16"] = to_numeric(df['16'], errors='coerce') 
                df["17"] = to_numeric(df['17'], errors='coerce') 
                df["18"] = to_numeric(df['18'], errors='coerce') 
                df["19"] = to_numeric(df['19'], errors='coerce') 
                df["20"] = to_numeric(df['20'], errors='coerce') 
                df["21"] = to_numeric(df['21'], errors='coerce') 
                df["22"] = to_numeric(df['22'], errors='coerce') 
                df["23"] = to_numeric(df['23'], errors='coerce') 
                df["24"] = to_numeric(df['24'], errors='coerce') 
                df["25"] = to_numeric(df['25'], errors='coerce') 
                df["26"] = to_numeric(df['26'], errors='coerce') 
                df["27"] = to_numeric(df['27'], errors='coerce') 
                df["28"] = to_numeric(df['28'], errors='coerce') 
                df["29"] = to_numeric(df['29'], errors='coerce') 
                df["30"] = to_numeric(df['30'], errors='coerce') 
                df["31"] = to_numeric(df['31'], errors='coerce') 
                df["Trung bình"] = to_numeric(df['Trung bình'], errors='coerce') 
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Adjust column width and format the header row
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            # Style the header row
            header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Create a date format for short date
            date_format = NamedStyle(name="Percentage", number_format="0.00%")
            if "Percentage" not in workbook.named_styles:
                workbook.add_named_style(date_format)
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        # Apply the date format to column L (assuming 'Ngày thực hiện' is in column 'L')
                        if cell.column_letter in ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                        'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ'] and cell.value is not None:
                            cell.number_format = '0.00%'
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the modified workbook to the output BytesIO object
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocao_hieusuat_tonghop_may_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_hieusuat_may_tonghop")

@app.route("/baocao_tienthuong_may", methods=["GET","POST"])
def baocao_tienthuong_may():
    if request.method == "GET":
        try:
            nam = request.args.get("nam")
            thang = request.args.get("thang")
            macongty = request.args.get("macongty")
            mst = request.args.get("mst")
            chuyen = request.args.get("chuyen")
            danhsach = lay_baocao_tienthuong_may(nam,thang,macongty,mst,chuyen)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 10
            total = len(danhsach)
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("baocao_tienthuong_may.html", danhsach=paginated_rows,pagination=pagination)
        except Exception as e:
            print(e)
            return render_template("baocao_tienthuong_may.html", danhsach=[])
    elif request.method == "POST":
        try:
            nam = request.form.get("nam")
            thang = request.form.get("thang")
            macongty = request.form.get("macongty")
            mst = request.form.get("mst")
            chuyen = request.form.get("chuyen")
            danhsach = lay_baocao_tienthuong_may(nam,thang,macongty,mst,chuyen)
            data = [{
                "Mã số thẻ": row[3],
                "Họ tên" : row[4],
                "Chuyền" : row[5],
                "SCP": row[6],
                "01" : row[7],
                "02" : row[8],
                "03" : row[9],
                "04" : row[10],
                "05" : row[11],
                "06" : row[12],
                "07" : row[13],
                "08" : row[14],
                "09" : row[15],
                "10" : row[16],
                "11" : row[17],
                "12" : row[18],
                "13" : row[19],
                "14" : row[20],
                "15" : row[21],
                "16" : row[22],
                "17" : row[23],
                "18" : row[24],
                "19" : row[25],
                "20" : row[26],
                "21" : row[27],
                "22" : row[28],
                "23" : row[29],
                "24" : row[30],
                "25" : row[31],
                "26" : row[32],
                "27" : row[33],
                "28" : row[34],
                "29" : row[35],
                "30" : row[36],
                "31" : row[37],
                "Tổng thưởng" : row[38]
            } for row in danhsach]
            df = DataFrame(data)
            if data:
                df['Mã số thẻ'] = to_numeric(df['Mã số thẻ'], errors='coerce')
                df["01"] = to_numeric(df['01'], errors='coerce') 
                df["02"] = to_numeric(df['02'], errors='coerce') 
                df["03"] = to_numeric(df['03'], errors='coerce') 
                df["04"] = to_numeric(df['04'], errors='coerce') 
                df["05"] = to_numeric(df['05'], errors='coerce') 
                df["06"] = to_numeric(df['06'], errors='coerce') 
                df["07"] = to_numeric(df['07'], errors='coerce') 
                df["08"] = to_numeric(df['08'], errors='coerce')
                df["09"] = to_numeric(df['09'], errors='coerce') 
                df["10"] = to_numeric(df['10'], errors='coerce') 
                df["11"] = to_numeric(df['11'], errors='coerce') 
                df["12"] = to_numeric(df['12'], errors='coerce') 
                df["13"] = to_numeric(df['13'], errors='coerce') 
                df["14"] = to_numeric(df['14'], errors='coerce') 
                df["15"] = to_numeric(df['15'], errors='coerce') 
                df["16"] = to_numeric(df['16'], errors='coerce') 
                df["17"] = to_numeric(df['17'], errors='coerce') 
                df["18"] = to_numeric(df['18'], errors='coerce') 
                df["19"] = to_numeric(df['19'], errors='coerce') 
                df["20"] = to_numeric(df['20'], errors='coerce') 
                df["21"] = to_numeric(df['21'], errors='coerce') 
                df["22"] = to_numeric(df['22'], errors='coerce') 
                df["23"] = to_numeric(df['23'], errors='coerce') 
                df["24"] = to_numeric(df['24'], errors='coerce') 
                df["25"] = to_numeric(df['25'], errors='coerce') 
                df["26"] = to_numeric(df['26'], errors='coerce') 
                df["27"] = to_numeric(df['27'], errors='coerce') 
                df["28"] = to_numeric(df['28'], errors='coerce') 
                df["29"] = to_numeric(df['29'], errors='coerce') 
                df["30"] = to_numeric(df['30'], errors='coerce') 
                df["31"] = to_numeric(df['31'], errors='coerce') 
                df["Tổng thưởng"] = to_numeric(df['Tổng thưởng'], errors='coerce') 
            output = BytesIO()
            with ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Adjust column width and format the header row
            output.seek(0)
            workbook = load_workbook(output)
            sheet = workbook.active

            # Style the header row
            header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Create a date format for short date
            date_format = NamedStyle(name="Percentage", number_format="0.00%")
            if "Percentage" not in workbook.named_styles:
                workbook.add_named_style(date_format)
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        # Apply the date format to column L (assuming 'Ngày thực hiện' is in column 'L')
                        if cell.column_letter in ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                        'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ'] and cell.value is not None:
                            cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the modified workbook to the output BytesIO object
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            time_stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
            # Trả file về cho client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=baocao_tienthuong_may_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response  
        except Exception as e:
            print(e)
            return redirect("/baocao_hieusuat_may_tonghop")
           
@app.route("/tailen_danhsach_dihotro", methods=["POST"])
def tailen_danhsach_dihotro():       
    if request.method == "POST":
        file = request.files["file"]
        thoigian = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
        filepath = f"tailen/dshotro_{thoigian}.xlsx"
        file.save(filepath)
        data = read_excel(filepath).to_dict(orient="records")
        for row in data:
            chuyendihotro = row["Chuyền đi hỗ trợ"]
            ngay = row["Ngày"].date()
            gio = row["Giờ"]
            sogio = row["Số giờ"]
            id = row["ID"]
            if capnhat_danhsach_dihotro(chuyendihotro,ngay,gio,sogio,id):
                flash("Đã cập nhật thành công !!!")
            else:
                flash("Đã cập nhật thất bại !!!")
        return redirect("/danhsach_dihotro")

@app.route("/sua_chuyendi_hotro", methods=["POST"])       
def sua_chuyendi_hotro():
    if request.method == "POST":
        chuyen = request.form.get("chuyen")
        id = request.form.get("id")
        doichuyendihotro(chuyen,id)
        return redirect("/danhsach_dihotro")     
        
@app.route("/sua_ngay_hotro", methods=["POST"])       
def sua_ngay_hotro():
    if request.method == "POST":
        ngay = request.form.get("ngay")
        id = request.form.get("id")
        doingaydihotro(ngay,id)
        return redirect("/danhsach_dihotro")    
        
@app.route("/sua_gio_hotro", methods=["POST"])       
def sua_gio_hotro():
    if request.method == "POST":
        gio = request.form.get("gio")
        id = request.form.get("id")
        doigiodihotro(gio,id)
        return redirect("/danhsach_dihotro")
        
@app.route("/sua_sogio_hotro", methods=["POST"])       
def sua_sogio_hotro():
    if request.method == "POST":
        sogio = request.form.get("sogio")
        id = request.form.get("id")
        doisogiodihotro(sogio,id)
        return redirect("/danhsach_dihotro")         

@app.route("/nhap_excel", methods=["GET"])
def nhap_excel():
    if request.method == "GET":
        try:    
            if "IED" in current_user.phongban:
                return redirect("/hieusuat_tnc")
            if "QAD" in current_user.phongban:
                return redirect("/ti_le_loi")
            return redirect("/")
        except Exception as e:
            print(e)
            return redirect("/")

if __name__ == "__main__":
    try:
        if sys.argv[1]=="1":
            while True:
                try:
                    serve(app, host="0.0.0.0", port=83, threads=8, _quiet=True)
                except subprocess.CalledProcessError as e:
                    print(f"Flask gap loi: {e}")
                    print("Đang khoi dong flask...")
                    time.sleep(1)  # Đợi một khoảng thời gian trước khi khởi động lại
                except Exception as e:
                    print(f"Loi khong xac dinh: {e}")
                    print("Đang khoi dong lai flask ...")
                    time.sleep(1)
    except:
        app.run(host="0.0.0.0", port=83, debug=True)
        
    